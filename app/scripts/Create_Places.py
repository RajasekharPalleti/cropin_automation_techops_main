"""
Creates places by taking 'Place name', 'Place type', 'Latitude', and 'Longitude' from an Excel file,
fetching the corresponding address from Google Maps API, and submitting it to the Cropin API.

Author: Rajasekhar Palleti

Inputs:
Excel file with 'Place name', 'Place type', 'Latitude', and 'Longitude' columns.
"""

import time
import requests
import json
import openpyxl

def get_address_data(session: requests.Session, lat: float, lng: float, google_api_key: str) -> dict:
    if not google_api_key:
        raise RuntimeError("Google API key is missing. Set Google API Key in config before running.")

    params = {
        "latlng": f"{lat},{lng}",
        "key": google_api_key,
        "result_type": "street_address|premise|sublocality|locality",
    }

    resp = session.get("https://maps.googleapis.com/maps/api/geocode/json", params=params, timeout=10)
    data = resp.json()

    if data.get("status") != "OK" or not data.get("results"):
        raise RuntimeError(f"Google API returned {data.get('status')} - could not fetch address for {lat},{lng}")

    result = data["results"][0]
    formatted_address = result.get("formatted_address", "")

    comps_raw = result.get("address_components", [])
    comps = {
        "country": "",
        "administrativeAreaLevel1": "",
        "administrativeAreaLevel2": "",
        "locality": "",
        "sublocalityLevel1": "",
        "sublocalityLevel2": "",
        "postalCode": "",
    }

    for c in comps_raw:
        t = c.get("types", [])
        if "country" in t:
            comps["country"] = c.get("long_name", "")
        elif "administrative_area_level_1" in t:
            comps["administrativeAreaLevel1"] = c.get("long_name", "")
        elif "administrative_area_level_2" in t:
            comps["administrativeAreaLevel2"] = c.get("long_name", "")
        elif "locality" in t:
            comps["locality"] = c.get("long_name", "")
        elif "sublocality_level_1" in t:
            comps["sublocalityLevel1"] = c.get("long_name", "")
        elif "sublocality_level_2" in t:
            comps["sublocalityLevel2"] = c.get("long_name", "")
        elif "postal_code" in t:
            comps["postalCode"] = c.get("long_name", "")

    geometry = result.get("geometry", {}).get("location", {})
    glat = geometry.get("lat", lat)
    glng = geometry.get("lng", lng)

    return {
        "country": comps["country"],
        "formattedAddress": formatted_address,
        "administrativeAreaLevel1": comps["administrativeAreaLevel1"],
        "administrativeAreaLevel2": comps["administrativeAreaLevel2"],
        "locality": comps["locality"],
        "sublocalityLevel1": comps["sublocalityLevel1"],
        "sublocalityLevel2": comps["sublocalityLevel2"],
        "landmark": "",
        "postalCode": comps["postalCode"],
        "houseNo": "",
        "buildingName": "",
        "placeId": result.get("place_id", ""),
        "latitude": glat,
        "longitude": glng,
    }


def build_payload(place_name: str, place_type: str, address_data: dict) -> dict:
    return {
        "data": None,
        "name": place_name,
        "type": place_type,
        "address": address_data,
        "latitude": address_data["latitude"],
        "longitude": address_data["longitude"],
    }


def run(input_excel_file, output_excel_file, config, log_callback=None):
    def log(msg):
        if log_callback:
            log_callback(msg)
        else:
            print(msg)

    api_url = config.get("base_api_url", "https://cloud.cropin.in/services/farm/api/place")
    google_api_key = config.get("x_api_key")  # Sourced from 'google_api' extended_config_type
    token = config.get("token")
    delay_time = float(config.get("delay_time", 1.0))

    if not token:
        log("Error: Authorization token missing.")
        return
        
    if not google_api_key:
        log("Error: Google API Key missing. Please provide it in the UI.")
        return

    headers = {
        "Authorization": f"Bearer {token}",
        "Content-Type": "application/json",
    }

    log(f"Reading input file: {input_excel_file}")
    try:
        wb = openpyxl.load_workbook(input_excel_file)
        sheet = wb.active
    except Exception as e:
        log(f"Failed to read Excel: {e}")
        return

    headers_row = [str(sheet.cell(1, c).value).strip() if sheet.cell(1, c).value else "" for c in range(1, sheet.max_column + 1)]
    next_col = sheet.max_column + 1
    
    # Append output columns if missing
    for col_name in ["Status", "Failure Reason", "Place ID"]:
        if col_name not in headers_row:
            sheet.cell(1, next_col, col_name)
            headers_row.append(col_name)
            next_col += 1

    try:
        name_col = headers_row.index("Place name") + 1
        type_col = headers_row.index("Place type") + 1
        lat_col = headers_row.index("Latitude") + 1
        lng_col = headers_row.index("Longitude") + 1
    except ValueError as e:
        log(f"Error: Missing required column in Excel. Expected exactly: 'Place name', 'Place type', 'Latitude', 'Longitude'.")
        return

    status_col = headers_row.index("Status") + 1
    reason_col = headers_row.index("Failure Reason") + 1
    pid_col = headers_row.index("Place ID") + 1

    session = requests.Session()
    max_row = sheet.max_row
    
    log(f"Starting place creation for {max_row - 1} rows...")

    for row in range(2, max_row + 1):
        place_name = sheet.cell(row, name_col).value
        place_type = sheet.cell(row, type_col).value
        lat = sheet.cell(row, lat_col).value
        lng = sheet.cell(row, lng_col).value

        # End of data check
        if not place_name and not place_type and lat is None and lng is None:
            continue

        if not place_name or not place_type or lat is None or lng is None:
            sheet.cell(row, status_col, "Skipped")
            sheet.cell(row, reason_col, "Missing required fields")
            continue
            
        try:
            lat_f = float(lat)
            lng_f = float(lng)
        except ValueError:
            sheet.cell(row, status_col, "Failed")
            sheet.cell(row, reason_col, "Latitude or Longitude is not a valid number")
            continue

        log(f"📍 Executing Row {row - 1} of {max_row - 1}: Creating place '{place_name}' at {lat_f},{lng_f}")

        try:
            # 1. Fetch Address
            address_data = get_address_data(session, lat_f, lng_f, google_api_key)
            
            # 2. Build Payload
            payload = build_payload(str(place_name), str(place_type), address_data)

            # 3. Create Place
            resp = session.post(
                api_url,
                headers=headers,
                data=json.dumps(payload),
                timeout=30,
            )
            
            if resp.status_code in (200, 201):
                try:
                    data = resp.json()
                    pid = data.get("id") or data.get("data", {}).get("id")
                    sheet.cell(row, status_col, "Success")
                    if pid:
                        sheet.cell(row, pid_col, str(pid))
                        log(f"   ✅ Successfully created place → ID: {pid}")
                    else:
                        log("   ✅ Success but failed to extract ID from JSON")
                except Exception:
                    sheet.cell(row, status_col, "Success (JSON Parse Failed)")
                    log("   ⚠️ Created but JSON parse failed")
            else:
                err_text = resp.text[:200]
                sheet.cell(row, status_col, "Failed")
                sheet.cell(row, reason_col, f"HTTP {resp.status_code}: {err_text}")
                log(f"   ❌ Failed ({resp.status_code}) → {err_text}")

        except Exception as e:
            sheet.cell(row, status_col, "Failed")
            sheet.cell(row, reason_col, str(e))
            log(f"   ❌ Error: {str(e)}")

        time.sleep(delay_time)

    wb.save(output_excel_file)
    log(f"✅ Output saved to: {output_excel_file}")
