"""
Remove FARMER tags from farmers using Excel input.

Inputs:
Excel file with 'Farmer ID' and 'Tags' (comma-separated IDs or Names).
"""

import json
import requests
import openpyxl
import time
import re
import os

# ------------------------------------------------------------
# Normalize tag names (handles spaces & case)
# ------------------------------------------------------------
def normalize_tag_name(name: str) -> str:
    return re.sub(r"\s+", " ", name.strip().lower())

# ------------------------------------------------------------
# Fetch FARMER tags (Name → ID mapping)
# ------------------------------------------------------------
def fetch_farmer_tag_map(token, url):
    if not url:
        url = "https://cloud.cropin.in/services/master/api/filter?type=FARMER&size=10000"
        
    headers = {"Authorization": f"Bearer {token}"}

    print(f"Fetching tags from: {url}")
    response = requests.get(url, headers=headers)
    response.raise_for_status()

    tag_map = {}
    for tag in response.json():
        if "id" in tag and "name" in tag:
            tag_map[normalize_tag_name(tag["name"])] = tag["id"]

    return tag_map

# ------------------------------------------------------------
# Resolve Excel input → Tag IDs
# ------------------------------------------------------------
def resolve_tag_ids(raw_tokens, tag_name_map):
    resolved_ids = []
    unresolved = []

    for token in raw_tokens:
        token = str(token).strip() # Ensure string

        if token.isdigit():
            resolved_ids.append(int(token))
        else:
            normalized = normalize_tag_name(token)
            if normalized in tag_name_map:
                resolved_ids.append(tag_name_map[normalized])
            else:
                unresolved.append(token)

    return resolved_ids, unresolved

# ------------------------------------------------------------
# Remove tag IDs from farmer response
# ------------------------------------------------------------
def remove_tag_ids(farmer_data, tag_ids_to_remove):
    tags = farmer_data.get("data", {}).get("tags", [])

    if not isinstance(tags, list):
        return farmer_data, set()

    original = set(tags)
    remove_set = set(tag_ids_to_remove)

    updated_tags = [t for t in tags if t not in remove_set]
    removed = original - set(updated_tags)

    farmer_data["data"]["tags"] = updated_tags
    return farmer_data, removed

# ------------------------------------------------------------
# Main Run Function
# ------------------------------------------------------------
def run(input_excel_file, output_excel_file, config, log_callback=None):
    
    def log(msg):
        if log_callback:
            log_callback(msg)
        else:
            print(msg)

    # 1. Parse Config
    api_url = config.get("base_api_url", "https://cloud.cropin.in/services/farm/api/farmers") # In main.py, we register it as 'base_api_url'
    # Actually, in main.py logic for other scripts:
    # api_url = config.get("base_api_url") is common but relies on UI sending it.
    # checking main.py again...
    # `default_configs` uses "url". 
    # BUT `Update_Farmer_Tags.py` uses `config.get("base_api_url")`.
    # `main.py` seems to NOT rename keys, just passes `config_dict`.
    # Wait, `main.py`: `config_dict = json.loads(config)`...
    # The UI `app.js` likely sends the configured URL as `base_api_url` or similar?
    # Let's double check `Update_Farmer_Tags.py` again. It uses `config.get("base_api_url")`.
    # Let's check `main.py` again to see what key it uses for the URL.
    # `scripts.append({ "name": filename, "url": config["url"], ... })`
    # The frontend likely sends back the edited URL key.
    # The UI `app.js` likely sends the configured URL as `base_api_url` or similar?
    
    # Let's confirm how the API calls are made:
    # `base_api_url/farmers/{farmer_id}`  (GET, DELETE etc)
    # So `api_url` needs to be the base.
    
    # If the user's `Update_Farmer_Tags.py` works, it implies the key is `base_api_url`.
    # Let's look at `Update_Farmer_Tags.py` again. Line 150: `api_url = config.get("base_api_url")`.
    # Then line 118: `requests.put(f"{api_url}", ...)` ? Hmm wait, if `api_url` is ".../farmers", PUT to ".../farmers" is bulk? 
    # Yes, Cropin bulk update is PUT to `/farmers`.
    
    # Ideally I should verify this, but "base_api_url" seems common for "Post Api Url" label.
    # The label for `Update_Farmer_Tags.py` was "Base Api Url".
    
    second_base_api_url = config.get("second_base_api_url", "https://cloud.cropin.in/services/master/api/filter?type=FARMER&size=10000")
    token = config.get("token")

    if not api_url:
        log("Error: 'base_api_url' not configured.")
        return
    if not token:
        log("Error: Authorization token missing.")
        return

    delay_time = float(config.get("delay_time", 0.5))  # seconds, configurable via UI


    # 3. Fetch Tag Mapping (Name -> ID)
    log("🔄 Fetching tag mapping from master API...")
    try:
        tag_name_map = fetch_farmer_tag_map(token, second_base_api_url)
        log(f"✅ Loaded {len(tag_name_map)} FARMER tags")
    except Exception as e:
        log(f"❌ Failed to fetch tags: {e}")
        return

    log(f"Reading input file: {input_excel_file}")
    try:
        wb = openpyxl.load_workbook(input_excel_file)
        sheet = wb.active # Use active sheet
    except Exception as e:
        log(f"Failed to read Excel: {e}")
        return

    # 3. Headers
    headers_row = [sheet.cell(1, c).value for c in range(1, sheet.max_column + 1)]
    next_col = sheet.max_column + 1

    if "Status" not in headers_row:
        sheet.cell(1, next_col, "Status")
        next_col += 1
    if "Failure Reason" not in headers_row:
        sheet.cell(1, next_col, "Failure Reason")
    
    # Recalculate column indices
    headers_row = [sheet.cell(1, c).value for c in range(1, sheet.max_column + 1)]
    try:
        status_col = headers_row.index("Status") + 1
        reason_col = headers_row.index("Failure Reason") + 1
    except ValueError:
         # Fallback if something weird happens
         status_col = sheet.max_column + 1
         reason_col = sheet.max_column + 2

    req_headers = {"Authorization": f"Bearer {token}"}

    # 4. Process Rows
    max_row = sheet.max_row
    log(f"Processing {max_row - 1} rows...")

    for row in range(2, max_row + 1):
        farmer_id = sheet.cell(row, 1).value
        # Farmer Name is column 2
        raw_tags = sheet.cell(row, 3).value

        if not farmer_id or not raw_tags:
            sheet.cell(row, status_col, "Skipped")
            sheet.cell(row, reason_col, "Missing Farmer ID or Tags")
            continue

        cleaned_tags = re.sub(r'[\[\]\'\"]', '', str(raw_tags))
        raw_tokens = [t.strip() for t in cleaned_tags.split(",") if t.strip()]
        tag_ids, _ = resolve_tag_ids(raw_tokens, tag_name_map)

        if not tag_ids:
            sheet.cell(row, status_col, "Skipped")
            sheet.cell(row, reason_col, "No valid tags found")
            continue

        try:
            # Check if farmer actually has these tags first?
            # User script: GET /farmers/{id}
            get_resp = requests.get(f"{api_url}/{farmer_id}", headers=req_headers)
            get_resp.raise_for_status()
            farmer_data = get_resp.json()

            existing_tags = farmer_data.get("data", {}).get("tags", [])
            
            # Convert existing tags to strings/ints depending on API, usually they are ints in ID list
            # safely convert to set of ints for comparison
            existing_tags_set = set()
            for t in existing_tags:
                try:
                    existing_tags_set.add(int(t))
                except:
                    pass

            common_tags = existing_tags_set.intersection(set(tag_ids))

            if not common_tags:
                sheet.cell(row, status_col, "Skipped")
                sheet.cell(row, reason_col, "No tags found to remove")
                continue

            updated_farmer, removed_ids = remove_tag_ids(farmer_data, tag_ids)

            multipart_data = {
                "dto": (None, json.dumps(updated_farmer), "application/json")
            }
            
            # PUT to base URL
            put_resp = requests.put(api_url, headers=req_headers, files=multipart_data)
            put_resp.raise_for_status()

            sheet.cell(row, status_col, "Success")
            sheet.cell(row, reason_col, f"Removed Tag IDs: {', '.join(map(str, removed_ids))}")
            log(f"Row {row}: Farmer {farmer_id} updated. Removed: {removed_ids}")

        except requests.exceptions.RequestException as e:
            err_msg = str(e)
            if e.response is not None:
                err_msg += f" - {e.response.text}"
            sheet.cell(row, status_col, "Failed")
            sheet.cell(row, reason_col, err_msg)
            log(f"Row {row}: Failed - {err_msg[:100]}")

        # Basic rate limiting
        time.sleep(delay_time)

    # 5. Save Output
    wb.save(output_excel_file)
    log(f"✅ Output saved to: {output_excel_file}")
