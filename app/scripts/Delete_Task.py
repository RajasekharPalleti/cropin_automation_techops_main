"""
Delete tasks via bulk API and verify deletion by fetching tasks for the given Croppable Area.

Batching Details:
Given an Excel sheet with 'Task_Id' and 'CA_Id' per row:
1. Groups all tasks by their associated CA_Id.
2. For each CA_Id group, deletes tasks in batches of maximum 50 Task IDs at a time.
3. Automatically maps the API completion status back to the exact rows where the task IDs were inputted.

Inputs:
Excel file with 'Task_Id' (a single task ID) and 'CA_Id' (the related Croppable Area ID) per row.
"""

import requests
import openpyxl
import time

def run(input_excel_file, output_excel_file, config, log_callback=None):
    def log(msg):
        if log_callback:
            log_callback(msg)
        print(msg)

    api_url_delete = config.get("base_api_url", "https://cloud.cropin.in/services/farm/api/tasks/bulk")
    api_url_verify = config.get("second_base_api_url", "https://cloud.cropin.in/services/farm/api/tasks/croppablearea")
    token = config.get("token")
    delay_time = float(config.get("delay_time", 1.0))

    if not token:
        log("Error: Authorization token missing.")
        return

    req_headers = {"Authorization": f"Bearer {token}"}

    log(f"Reading input file: {input_excel_file}")
    try:
        wb = openpyxl.load_workbook(input_excel_file)
        sheet = wb.active
    except Exception as e:
        log(f"Failed to read Excel: {e}")
        return

    headers_row = [sheet.cell(1, c).value for c in range(1, sheet.max_column + 1)]
    next_col = sheet.max_column + 1
    
    if "Status" not in headers_row:
        sheet.cell(1, next_col, "Status")
        next_col += 1
    if "Remarks" not in headers_row:
        sheet.cell(1, next_col, "Remarks")
        
    headers_row = [sheet.cell(1, c).value for c in range(1, sheet.max_column + 1)]
    try:
        status_col = headers_row.index("Status") + 1
        reason_col = headers_row.index("Remarks") + 1
    except ValueError:
        status_col = sheet.max_column + 1
        reason_col = sheet.max_column + 2

    # Identify Task_Id and CA_Id columns
    try:
        task_id_col = headers_row.index("Task_Id") + 1
    except ValueError:
        try:
            task_id_col = headers_row.index("Task_Ids") + 1
        except ValueError:
            task_id_col = 1
            log("Warning: Could not find 'Task_Id' header. Falling back to Column 1.")
            
    try:
        ca_id_col = headers_row.index("CA_Id") + 1
    except ValueError:
        ca_id_col = 2
        log("Warning: Could not find 'CA_Id' header. Falling back to Column 2.")

    max_row = sheet.max_row
    log(f"Scanning {max_row - 1} rows to group by CA_Id...")

    ca_groups = {}
    skipped_count = 0

    for row in range(2, max_row + 1):
        task_id_raw = sheet.cell(row, task_id_col).value
        ca_id_raw = sheet.cell(row, ca_id_col).value

        # End of data check
        if task_id_raw is None and ca_id_raw is None:
            break

        if not task_id_raw or not ca_id_raw:
            sheet.cell(row, status_col, "Skipped")
            sheet.cell(row, reason_col, "Missing Task_Id or CA_Id")
            skipped_count += 1
            continue

        # Handle numeric values and convert to string properly
        if isinstance(task_id_raw, float):
            task_id_str = str(int(task_id_raw)).strip()
        else:
            task_id_str = str(task_id_raw).strip()
            
        if isinstance(ca_id_raw, float):
            ca_id_str = str(int(ca_id_raw)).strip()
        else:
            ca_id_str = str(ca_id_raw).strip()
            
        if ca_id_str not in ca_groups:
            ca_groups[ca_id_str] = []
        ca_groups[ca_id_str].append((row, task_id_str))

    log(f"Found {len(ca_groups)} unique CA IDs. (Skipped {skipped_count} invalid rows)")

    # Execute by groups and batches
    for ca_id_str, items in ca_groups.items():
        batch_size = 50
        total_tasks = len(items)
        log(f"--- Processing CA_Id: {ca_id_str} ({total_tasks} tasks) ---")

        for i in range(0, total_tasks, batch_size):
            batch = items[i:i + batch_size]
            batch_task_ids = [item[1] for item in batch]
            clean_task_ids_str = ",".join(batch_task_ids)
            batch_num = (i // batch_size) + 1
            total_batches = (total_tasks + batch_size - 1) // batch_size
            
            log(f"[CA: {ca_id_str}] Batch {batch_num}/{total_batches}: Deleting {len(batch)} tasks...")
            
            try:
                # 1. Delete Tasks Request
                delete_url = f"{api_url_delete}?ids={clean_task_ids_str}"
                del_resp = requests.delete(delete_url, headers=req_headers)
                del_resp.raise_for_status()
                
                try:
                    del_json = del_resp.json()
                    deletable = del_json.get("deletable", 0)
                    non_deletable = del_json.get("nonDeletable", 0)
                    log(f"[CA: {ca_id_str}] Batch {batch_num}: Delete call success (Deletable: {deletable}, Non-Deletable: {non_deletable})")
                except Exception:
                    log(f"[CA: {ca_id_str}] Batch {batch_num}: Delete API responded successfully.")
                
                # Wait for 1 sec
                time.sleep(1)
                
                # 2. Verify Deletion Status Request
                log(f"[CA: {ca_id_str}] Batch {batch_num}: Verifying deletion via Croppable Area API...")
                verify_url = f"{api_url_verify}/{ca_id_str}?sort=lastModifiedDate,desc"
                ver_resp = requests.get(verify_url, headers=req_headers)
                ver_resp.raise_for_status()
                
                task_dtos = ver_resp.json()
                
                # The API returns a direct JSON array of tasks: [{"id": 3832313, ...}, ...]
                if isinstance(task_dtos, list):
                    tasks_list = task_dtos
                else:
                    # Fallback just in case the API changes its response format slightly 
                    # without breaking the script completely.
                    tasks_list = [task_dtos] if not isinstance(task_dtos, dict) else task_dtos.get("content", task_dtos.get("data", task_dtos.get("taskDTOList", [task_dtos])))
                    
                active_task_ids_in_ca = set(str(t.get("id", "")) for t in tasks_list if isinstance(t, dict) and "id" in t)
                
                not_deleted_count = 0
                for row_num, tid in batch:
                    if tid in active_task_ids_in_ca:
                        sheet.cell(row_num, status_col, "Not Deleted")
                        sheet.cell(row_num, reason_col, f"Task ID {tid} was not deleted from CA {ca_id_str}")
                        not_deleted_count += 1
                    else:
                        sheet.cell(row_num, status_col, "Deleted")
                        sheet.cell(row_num, reason_col, "Task successfully deleted")
                
                if not_deleted_count > 0:
                    log(f"[CA: {ca_id_str}] Batch {batch_num}: Complete. {not_deleted_count} tasks were not deleted.")
                else:
                    log(f"[CA: {ca_id_str}] Batch {batch_num}: Complete. All tasks deleted successfully.")
                    
            except requests.exceptions.RequestException as e:
                err_msg = str(e)
                if hasattr(e, 'response') and e.response is not None:
                    try:
                        err_msg += f" - {e.response.text}"
                    except:
                        pass
                for row_num, tid in batch:
                    sheet.cell(row_num, status_col, "Failed")
                    sheet.cell(row_num, reason_col, err_msg)
                log(f"[CA: {ca_id_str}] Batch {batch_num}: Failed - {err_msg[:100]}")

            # Throttle delay
            time.sleep(delay_time)

    # Save Output
    wb.save(output_excel_file)
    log(f"✅ Output saved to: {output_excel_file}")
