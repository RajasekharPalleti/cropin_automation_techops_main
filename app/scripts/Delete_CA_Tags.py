"""
Remove CA tags from croppable areas using Excel input.

Inputs:
Excel file with 'CA ID' and 'Tags' (comma-separated IDs or Names).
"""

import json
import requests
import openpyxl
import time
import re

# ------------------------------------------------------------
# Normalize tag names (handles spaces & case)
# ------------------------------------------------------------
def normalize_tag_name(name: str) -> str:
    return re.sub(r"\s+", " ", name.strip().lower())

# ------------------------------------------------------------
# Fetch CA tags (Name → ID mapping)
# ------------------------------------------------------------
def fetch_ca_tag_map(token, url):
    if not url:
        url = "https://cloud.cropin.in/services/master/api/filter?type=CA&size=10000"
        
    headers = {"Authorization": f"Bearer {token}"}

    print(f"Fetching tags from: {url}")
    response = requests.get(url, headers=headers)
    response.raise_for_status()

    tag_map = {}
    for tag in response.json():
        if "id" in tag and "name" in tag:
            tag_map[normalize_tag_name(tag["name"])] = tag["id"]

    return tag_map

# ------------------------------------------------------------
# Resolve Excel input → Tag IDs
# ------------------------------------------------------------
def resolve_tag_ids(raw_tokens, tag_name_map):
    resolved_ids = []
    unresolved = []

    for token in raw_tokens:
        token = str(token).strip() # Ensure string

        if token.isdigit():
            resolved_ids.append(int(token))
        else:
            normalized = normalize_tag_name(token)
            if normalized in tag_name_map:
                resolved_ids.append(tag_name_map[normalized])
            else:
                unresolved.append(token)

    return resolved_ids, unresolved

# ------------------------------------------------------------
# Main Run Function
# ------------------------------------------------------------
def run(input_excel_file, output_excel_file, config, log_callback=None):
    
    def log(msg):
        if log_callback:
            log_callback(msg)
        else:
            print(msg)

    # 1. Parse Config
    api_url = config.get("url") or config.get("post_api_url")
    tag_api_url = config.get("url2") or config.get("tag_api_url")
    token = config.get("token")

    if not api_url:
        log("Error: API URL not configured.")
        return
    if not token:
        log("Error: Authorization token missing.")
        return

    delay_time = float(config.get("delay_time", 0.5))  # seconds, configurable via UI

    # 2. Setup
    log(f"Fetching Tag Map from Master API...")
    try:
        tag_name_map = fetch_ca_tag_map(token, tag_api_url)
        log(f"✅ Loaded {len(tag_name_map)} CA tags")
    except Exception as e:
        log(f"❌ Failed to fetch tags: {e}")
        return

    log(f"Reading input file: {input_excel_file}")
    try:
        wb = openpyxl.load_workbook(input_excel_file)
        sheet = wb.active # Use active sheet
    except Exception as e:
        log(f"Failed to read Excel: {e}")
        return

    # 3. Headers
    headers_row = [sheet.cell(1, c).value for c in range(1, sheet.max_column + 1)]
    next_col = sheet.max_column + 1

    if "Status" not in headers_row:
        sheet.cell(1, next_col, "Status")
        next_col += 1
    if "Failure Reason" not in headers_row:
        sheet.cell(1, next_col, "Failure Reason")
    
    # Recalculate column indices
    headers_row = [sheet.cell(1, c).value for c in range(1, sheet.max_column + 1)]
    try:
        status_col = headers_row.index("Status") + 1
        reason_col = headers_row.index("Failure Reason") + 1
    except ValueError:
         status_col = sheet.max_column + 1
         reason_col = sheet.max_column + 2

    req_headers = {"Authorization": f"Bearer {token}"}

    # 4. Process Rows
    max_row = sheet.max_row
    log(f"Processing {max_row - 1} rows...")

    for row in range(2, max_row + 1):
        ca_id = sheet.cell(row, 1).value
        raw_tags = sheet.cell(row, 2).value

        # End of data check
        if ca_id is None and raw_tags is None:
            break

        if not ca_id or not raw_tags:
            sheet.cell(row, status_col, "Skipped")
            sheet.cell(row, reason_col, "Missing CA ID or Tags")
            continue

        ca_id = str(ca_id).strip()
        raw_tokens = [t for t in str(raw_tags).split(",") if t.strip()]
        tag_ids, unresolved = resolve_tag_ids(raw_tokens, tag_name_map)

        if unresolved:
            log(f"Warning mapping tags for CA {ca_id}: Could not resolve '{', '.join(unresolved)}'")

        if not tag_ids:
            sheet.cell(row, status_col, "Skipped")
            sheet.cell(row, reason_col, "No valid tags found matching system records")
            continue

        try:
            # First, fetch the existing CA
            get_resp = requests.get(f"{api_url}/{ca_id}", headers=req_headers)
            get_resp.raise_for_status()
            ca_data = get_resp.json()

            # Ensure data wrapper exists
            if "data" not in ca_data:
                ca_data["data"] = {}
                
            existing_tags = ca_data["data"].get("tags", [])
            
            # The API usually expects list of dicts: [{"id": 123}, {"id": 124}]
            # Or standard int list [123, 124]? Let's check CA API standard. Most APIs use list of ints or dicts.
            # Using dict format based on AddTagsWithNewAPI which does:
            # tags_list = [{"id": tid} for tid in common_tags]
            # Wait, `Delete_Farmer_Tags` just replaced with a list of ints.
            # Let's clean the existing tags.
            
            existing_tag_ids = set()
            for t in existing_tags:
                try:
                    existing_tag_ids.add(int(t))
                except:
                    pass

            to_remove_set = set(tag_ids)
            intersection = existing_tag_ids.intersection(to_remove_set)

            if not intersection:
                sheet.cell(row, status_col, "Skipped")
                sheet.cell(row, reason_col, "No specified tags exist on this CA to remove")
                continue

            # Remove the specified tags from the active set
            final_tag_ids = existing_tag_ids - to_remove_set
            
            # Format as expected by Cropin CA PUT API (list of ints)
            ca_data["data"]["tags"] = list(final_tag_ids)

            multipart_data = {
                "dto": (None, json.dumps(ca_data), "application/json")
            }
            
            # Update the CA
            put_resp = requests.put(api_url, headers=req_headers, files=multipart_data)
            put_resp.raise_for_status()

            sheet.cell(row, status_col, "Success")
            sheet.cell(row, reason_col, f"Removed Tag IDs: {', '.join(map(str, intersection))}")
            log(f"Row {row}: CA {ca_id} updated. Removed: {intersection}")

        except requests.exceptions.RequestException as e:
            err_msg = str(e)
            if e.response is not None:
                err_msg += f" - {e.response.text}"
            sheet.cell(row, status_col, "Failed")
            sheet.cell(row, reason_col, err_msg)
            log(f"Row {row}: Failed - {err_msg[:100]}")

        # Throttle delay
        time.sleep(delay_time)

    # 5. Save Output
    wb.save(output_excel_file)
    log(f"✅ Output saved to: {output_excel_file}")
