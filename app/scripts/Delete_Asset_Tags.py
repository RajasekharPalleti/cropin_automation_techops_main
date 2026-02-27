"""
Remove ASSET tags from assets using Excel input.

Inputs:
Excel file with 'Asset ID' and 'Tags' (comma-separated IDs or Names).
"""

import json
import requests
import openpyxl
import time
import re
import os

# ------------------------------------------------------------
# Normalize tag names (handles spaces & case)
# ------------------------------------------------------------
def normalize_tag_name(name: str) -> str:
    """
    Normalize tag names for consistent comparison.
    """
    return re.sub(r"\s+", " ", name.strip().lower())


# ------------------------------------------------------------
# Fetch ASSET tags (Name → ID mapping)
# ------------------------------------------------------------
def fetch_asset_tag_map(token, url):
    """
    Fetch all ASSET tags from the Master Tag API.
    """
    if not url:
        url = "https://cloud.cropin.in/services/master/api/filter?type=ASSET&size=10000"
        
    headers = {"Authorization": f"Bearer {token}"}

    print(f"Fetching tags from: {url}")
    response = requests.get(url, headers=headers)
    response.raise_for_status()

    tag_map = {}
    for tag in response.json():
        if "id" in tag and "name" in tag:
            tag_map[normalize_tag_name(tag["name"])] = tag["id"]

    print(f"✅ Loaded {len(tag_map)} ASSET tags")
    return tag_map


# ------------------------------------------------------------
# Resolve Excel input → Tag IDs
# ------------------------------------------------------------
def resolve_tag_ids(raw_tokens, tag_name_map):
    """
    Resolve tag IDs from Excel input (IDs / Names / Mixed).
    """
    resolved_ids = []
    unresolved = []

    for token in raw_tokens:
        token = str(token).strip()

        if token.isdigit():
            resolved_ids.append(int(token))
        else:
            normalized = normalize_tag_name(token)
            if normalized in tag_name_map:
                resolved_ids.append(tag_name_map[normalized])
            else:
                unresolved.append(token)

    return resolved_ids, unresolved


# ------------------------------------------------------------
# Remove tag IDs from asset response
# ------------------------------------------------------------
def remove_tag_ids(asset_data, tag_ids_to_remove):
    """
    Remove specified tag IDs from asset response.
    """
    tags = asset_data.get("data", {}).get("tags", [])

    if not isinstance(tags, list):
        return asset_data, set()

    original = set(tags)
    remove_set = set(tag_ids_to_remove)

    updated_tags = [t for t in tags if t not in remove_set]
    removed = original - set(updated_tags)

    asset_data["data"]["tags"] = updated_tags
    return asset_data, removed


# ------------------------------------------------------------
# Main Run Function
# ------------------------------------------------------------
def run(input_excel_file, output_excel_file, config, log_callback=None):
    
    def log(msg):
        if log_callback:
            log_callback(msg)
        else:
            print(msg)

    api_url = config.get("url") or config.get("post_api_url")
    tag_api_url = config.get("url2") or config.get("tag_api_url")
    token = config.get("token")

    if not api_url:
        log("Error: API URL not configured.")
        return
    if not token:
        log("Error: Authorization token missing.")
        return

    delay_time = float(config.get("delay_time", 0.5))  # seconds, configurable via UI

    # 2. Setup
    log(f"Fetching Tag Map from Master API...")
    try:
        tag_name_map = fetch_asset_tag_map(token, tag_api_url)
    except Exception as e:
        log(f"❌ Failed to fetch tags: {e}")
        return

    log(f"Reading input file: {input_excel_file}")
    try:
        wb = openpyxl.load_workbook(input_excel_file)
        sheet = wb.active # Use active sheet
    except Exception as e:
        log(f"Failed to read Excel: {e}")
        return

    # 3. Headers
    headers_row = [sheet.cell(1, c).value for c in range(1, sheet.max_column + 1)]
    next_col = sheet.max_column + 1

    if "Status" not in headers_row:
        sheet.cell(1, next_col, "Status")
        next_col += 1

    if "Failure Reason" not in headers_row:
        sheet.cell(1, next_col, "Failure Reason")
    
    # Recalculate column indices
    headers_row = [sheet.cell(1, c).value for c in range(1, sheet.max_column + 1)]
    try:
        status_col = headers_row.index("Status") + 1
        reason_col = headers_row.index("Failure Reason") + 1
    except ValueError:
         status_col = sheet.max_column + 1
         reason_col = sheet.max_column + 2

    req_headers = {"Authorization": f"Bearer {token}"}

    # 4. Process Rows
    max_row = sheet.max_row
    log(f"Processing {max_row - 1} rows...")

    for row in range(2, max_row + 1):
        asset_id = sheet.cell(row, 1).value
        # Asset Name is column 2
        raw_tags = sheet.cell(row, 3).value

        if not asset_id or not raw_tags:
            sheet.cell(row, status_col, "Skipped")
            sheet.cell(row, reason_col, "Missing Asset ID or Tags")
            continue

        cleaned_tags = re.sub(r'[\[\]\'\"]', '', str(raw_tags))
        raw_tokens = [t.strip() for t in cleaned_tags.split(",") if t.strip()]
        tag_ids, _ = resolve_tag_ids(raw_tokens, tag_name_map)

        if not tag_ids:
            sheet.cell(row, status_col, "Skipped")
            sheet.cell(row, reason_col, "No valid tags found")
            continue

        try:
            # GET Asset
            get_resp = requests.get(f"{api_url}/{asset_id}", headers=req_headers)
            get_resp.raise_for_status()
            asset_data = get_resp.json()

            existing_tags = asset_data.get("data", {}).get("tags", [])
            
            # Convert existing to set of ints if possible
            existing_tags_set = set()
            for t in existing_tags:
                try:
                    existing_tags_set.add(int(t))
                except:
                    pass

            common_tags = existing_tags_set.intersection(set(tag_ids))

            if not common_tags:
                sheet.cell(row, status_col, "Skipped")
                sheet.cell(row, reason_col, "No tags found to remove")
                continue

            updated_asset, removed_ids = remove_tag_ids(asset_data, tag_ids)

            multipart_data = {
                "dto": (None, json.dumps(updated_asset), "application/json")
            }

            time.sleep(delay_time)

            # PUT Asset update
            put_resp = requests.put(api_url, headers=req_headers, files=multipart_data)
            put_resp.raise_for_status()

            sheet.cell(row, status_col, "Success")
            sheet.cell(
                row,
                reason_col,
                f"Removed Tag IDs: {', '.join(map(str, removed_ids))}"
            )
            log(f"Row {row}: Asset {asset_id} updated. Removed: {removed_ids}")

        except requests.exceptions.RequestException as e:
            err_msg = str(e)
            if e.response is not None:
                err_msg += f" - {e.response.text}"
            sheet.cell(row, status_col, "Failed")
            sheet.cell(row, reason_col, err_msg)
            log(f"Row {row}: Failed - {err_msg[:100]}")

        time.sleep(delay_time)

    # Save to OUTPUT file
    wb.save(output_excel_file)
    log(f"📁 Output saved to: {output_excel_file}")
